#!/usr/bin/env python3
"""
Import KPI data from Excel files into Supabase.

Usage:
    cd /path/to/wetpaint-human-connections
    pip install openpyxl supabase python-dotenv --break-system-packages
    python scripts/import_kpi.py

Requires SUPABASE_SERVICE_ROLE_KEY in .env.local (alongside the existing
NEXT_PUBLIC_SUPABASE_URL). The script is idempotent — safe to re-run.
"""

import os
import re
import sys
from pathlib import Path

from typing import Optional

from dotenv import load_dotenv
import openpyxl
from supabase import create_client, Client

# ---------------------------------------------------------------------------
# Paths & constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent.parent
KPI_DIR = BASE_DIR / "supplied-by-hr" / "KPIS 2026"
ENV_FILE = BASE_DIR / ".env.local"

# These scorer first-names always resolve to scorer_id = NULL (HR/management).
# Ujala is explicitly required by spec; Petra is listed as management scorer;
# "manager" / "line" cover "Manager - Score" / "Line Manager - Score" cells.
MANAGEMENT_SCORERS = {"ujala", "petra", "manager", "line"}

# Section definitions in display/insertion order
SECTIONS_META = [
    {"title": "KPI FOCUS AREAS: PERSONAL & DEPARTMENT", "type": "invitee", "position": 1},
    {"title": "OUR VALUES : TRIBE CONTRACT",            "type": "invitee", "position": 2},
    {"title": "HR AREAS: TRIBE CONTRACT",               "type": "hr",      "position": 3},
]

# Excel layout: header row, item rows, total row per section index
SECTION_ROWS = {
    0: {"header": 14, "items": [15, 16, 17]},
    1: {"header": 21, "items": [22, 23, 24, 25]},
    2: {"header": 28, "items": [29, 30, 31, 32, 33, 34, 35]},
}


# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------
def load_env():
    load_dotenv(ENV_FILE)
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print(
            "ERROR: .env.local must contain both NEXT_PUBLIC_SUPABASE_URL "
            "and SUPABASE_SERVICE_ROLE_KEY"
        )
        sys.exit(1)
    return url, key


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
def parse_scorer_name(cell_value):
    """
    Extract the first name token from a scorer header cell.
    e.g. 'Kerry - Score' → 'Kerry',  'Petra Score' → 'Petra',
         'Ujala - HR Manager Score' → 'Ujala',  'Manager - Score' → 'Manager'
    """
    if not cell_value:
        return None
    text = str(cell_value).strip()
    # Split on ' - ' (with optional spaces) or whitespace, take first token
    token = re.split(r"\s*[-–]\s*|\s+", text)[0]
    return token.strip() or None


def parse_employee_number(b4_value):
    """Extract WPxxxxxx code from e.g. 'Darren Maxwell (WP068DL)'."""
    if not b4_value:
        return None
    m = re.search(r"\(([A-Z0-9]+)\)", str(b4_value))
    return m.group(1).upper() if m else None


def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Employee lookups
# ---------------------------------------------------------------------------
def build_employee_lookups(supabase: Client):
    """
    Returns:
      by_number: {EMPLOYEE_NUMBER: employee_row}  — keyed by upper-case code
      by_first:  {first_word_lower: employee_row} — keyed by first word of first_name
    """
    res = supabase.table("employees").select("id,first_name,last_name,employee_number").execute()
    by_number: dict = {}
    by_first: dict = {}
    for emp in res.data:
        num = (emp.get("employee_number") or "").upper().strip()
        if num:
            by_number[num] = emp
        first_name = emp.get("first_name") or ""
        first_word = first_name.split()[0].lower() if first_name.split() else ""
        if first_word and first_word not in by_first:
            by_first[first_word] = emp
    return by_number, by_first


def resolve_employee(folder_name: str, b4_value, by_number: dict, by_first: dict):
    # Priority 1: match by employee_number extracted from cell B4
    emp_num = parse_employee_number(b4_value)
    if emp_num and emp_num in by_number:
        return by_number[emp_num]
    # Priority 2: match folder name against first word of first_name
    return by_first.get(folder_name.lower())


def resolve_scorer(name, by_first: dict):
    """Return employee_id, or None for management scorers / unmatched names."""
    if not name:
        return None
    key = name.lower()
    if key in MANAGEMENT_SCORERS:
        return None
    emp = by_first.get(key)
    return emp["id"] if emp else None


# ---------------------------------------------------------------------------
# Upsert helpers (all idempotent)
# ---------------------------------------------------------------------------
def upsert_section(supabase: Client, meta: dict) -> str:
    res = supabase.table("kpi_template_sections").select("id").eq("title", meta["title"]).execute()
    if res.data:
        return res.data[0]["id"]
    ins = supabase.table("kpi_template_sections").insert({
        "title":     meta["title"],
        "type":      meta["type"],
        "position":  meta["position"],
        "is_active": True,
    }).execute()
    return ins.data[0]["id"]


def upsert_item(supabase: Client, section_id: str, title: str, description: str, position: int) -> str:
    res = (
        supabase.table("kpi_template_items")
        .select("id")
        .eq("section_id", section_id)
        .eq("title", title)
        .execute()
    )
    if res.data:
        return res.data[0]["id"]
    ins = supabase.table("kpi_template_items").insert({
        "section_id":  section_id,
        "title":       title,
        "description": description,
        "min_score":   0,
        "max_score":   10,
        "position":    position,
        "is_active":   True,
    }).execute()
    return ins.data[0]["id"]


def upsert_review(supabase: Client, employee_id: str, period: str) -> str:
    res = (
        supabase.table("kpi_reviews")
        .select("id")
        .eq("employee_id", employee_id)
        .eq("period", period)
        .execute()
    )
    if res.data:
        return res.data[0]["id"]
    ins = supabase.table("kpi_reviews").insert({
        "employee_id": employee_id,
        "period":      period,
        "title":       f"KPI Review – {period}",
        "status":      "active",
        "deadline":    None,
    }).execute()
    return ins.data[0]["id"]


def upsert_invitee(supabase: Client, review_id: str, invitee_id: str) -> None:
    supabase.table("kpi_review_invitees").upsert(
        {"review_id": review_id, "invitee_id": invitee_id, "status": "completed"},
        on_conflict="review_id,invitee_id",
    ).execute()


def upsert_score(supabase: Client, review_id: str, item_id: str, scorer_id, score: int) -> None:
    if scorer_id is not None:
        # DB has a unique constraint on (review_id, item_id, scorer_id)
        supabase.table("kpi_scores").upsert(
            {"review_id": review_id, "item_id": item_id, "scorer_id": scorer_id, "score": score, "comments": None},
            on_conflict="review_id,item_id,scorer_id",
        ).execute()
    else:
        # NULL != NULL in Postgres unique constraints — must handle manually.
        # When two NULL-scorer entries exist per item (e.g. Ujala + Manager in Cat 3),
        # the second call updates the row inserted by the first (last write wins).
        res = (
            supabase.table("kpi_scores")
            .select("id")
            .eq("review_id", review_id)
            .eq("item_id", item_id)
            .is_("scorer_id", "null")
            .execute()
        )
        if res.data:
            supabase.table("kpi_scores").update({"score": score, "comments": None}).eq("id", res.data[0]["id"]).execute()
        else:
            supabase.table("kpi_scores").insert(
                {"review_id": review_id, "item_id": item_id, "scorer_id": None, "score": score, "comments": None}
            ).execute()


# ---------------------------------------------------------------------------
# Per-file processing
# ---------------------------------------------------------------------------
def process_file(
    path: Path,
    folder_name: str,
    supabase: Client,
    section_ids,
    by_number: dict,
    by_first: dict,
    shared_item_ids: dict,  # mutable cache: {(sec_idx, pos) → item_id} for Cat 2 & 3
):
    """
    Parse one Excel file and write data to Supabase.
    Returns (items_upserted, scores_upserted).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    b4 = ws["B4"].value
    period = str(ws["B7"].value or "").strip()

    emp = resolve_employee(folder_name, b4, by_number, by_first)
    if emp is None:
        print(f"  ✗ {folder_name} ({b4}) — employee not found in Supabase, skipped")
        return 0, 0

    emp_id = emp["id"]
    review_id = upsert_review(supabase, emp_id, period)

    items_count = 0
    scores_count = 0
    invitees_added: set[str] = set()

    for sec_idx, rows in SECTION_ROWS.items():
        section_id = section_ids[sec_idx]
        header_row = rows["header"]

        c_name = parse_scorer_name(ws.cell(row=header_row, column=3).value)
        g_name = parse_scorer_name(ws.cell(row=header_row, column=7).value)
        c_scorer_id = resolve_scorer(c_name, by_first)
        g_scorer_id = resolve_scorer(g_name, by_first)

        # Register non-management scorers as invitees (exclude the employee themselves)
        for sid in (c_scorer_id, g_scorer_id):
            if sid and sid != emp_id and sid not in invitees_added:
                invitees_added.add(sid)
                upsert_invitee(supabase, review_id, sid)

        for pos, item_row in enumerate(rows["items"]):
            title = str(ws.cell(row=item_row, column=1).value or "").strip()
            description = str(ws.cell(row=item_row, column=2).value or "").strip()
            c_score = safe_int(ws.cell(row=item_row, column=3).value)
            g_score = safe_int(ws.cell(row=item_row, column=7).value)

            if not title:
                continue

            # Cat 1 items are personalised per employee → always upsert individually.
            # Cat 2 & 3 items are shared → cache after first insert to avoid redundant SELECTs.
            if sec_idx == 0:
                item_id = upsert_item(supabase, section_id, title, description, pos + 1)
            else:
                cache_key = (sec_idx, pos)
                if cache_key in shared_item_ids:
                    item_id = shared_item_ids[cache_key]
                else:
                    item_id = upsert_item(supabase, section_id, title, description, pos + 1)
                    shared_item_ids[cache_key] = item_id

            items_count += 1

            if c_score is not None:
                upsert_score(supabase, review_id, item_id, c_scorer_id, c_score)
                scores_count += 1
            if g_score is not None:
                upsert_score(supabase, review_id, item_id, g_scorer_id, g_score)
                scores_count += 1

    display = str(b4 or folder_name)
    print(f"  ✓ {display} — {period} — {items_count} items, {scores_count} scores upserted")
    return items_count, scores_count


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main() -> None:
    url, key = load_env()
    supabase: Client = create_client(url, key)

    # ── Step 1: template sections ──────────────────────────────────────────
    print("Step 1: Upserting template sections...")
    section_ids = []
    for meta in SECTIONS_META:
        sid = upsert_section(supabase, meta)
        section_ids.append(sid)
        print(f"  '{meta['title']}' → {sid}")

    # ── Step 2: employee files ─────────────────────────────────────────────
    print("\nStep 2: Processing employee KPI files...")
    by_number, by_first = build_employee_lookups(supabase)

    shared_item_ids: dict = {}
    total_items = total_scores = processed = skipped = 0

    for folder in sorted(f for f in KPI_DIR.iterdir() if f.is_dir()):
        name = folder.name
        xlsx = folder / f"{name} 2026 KPI Doc.xlsx"
        if not xlsx.exists():
            print(f"  ✗ {name} — {xlsx.name} not found, skipped")
            skipped += 1
            continue
        try:
            items, scores = process_file(
                xlsx, name, supabase, section_ids, by_number, by_first, shared_item_ids
            )
            if items == 0 and scores == 0:
                skipped += 1
            else:
                total_items += items
                total_scores += scores
                processed += 1
        except Exception as exc:
            print(f"  ✗ {name} — ERROR: {exc}")
            skipped += 1

    print(f"\nDone. {processed} reviews written, {total_scores} scores upserted, {skipped} files skipped.")


if __name__ == "__main__":
    main()
